"""extract_requirements.py — v2 deterministic requirement-row extractor.

Reads source Excel files directly. Emits one JSON record per requirement row
per expert. requirement_text and evidence are verbatim source cell values with
leading/trailing whitespace trimmed (via _s) — internal text is untouched;
trimming keeps the content-only dedup key from splitting on stray end-whitespace.

Usage:
    python extract_requirements.py <file1> [<file2> ...]
    python extract_requirements.py --all
"""

import json
import re
import sys
import unicodedata
from datetime import date, datetime
from pathlib import Path

import openpyxl

from config import SYNC_ROOT
from enumerate_candidates import enumerate_candidates
from dedupe import deduplicate

OUTPUT_FILE = Path(__file__).parent / "extraction_batch.json"

# ── sheet filtering ───────────────────────────────────────────────────────────

# Helper (non-expert) sheets. "ohje"/"ohjeet"/"pisteet" instruction & score sheets are matched
# followed by whitespace or end-of-name, so numbered/suffixed variants are caught ("Ohjeet 2",
# "Pisteet yhteensä") WITHOUT swallowing real role sheets that merely start with those letters
# (e.g. "Ohjelmoija" — a programmer). "data" is matched exact or with an explicit separator so
# content sheets like "Datakonversio" are not skipped.
_HELPER_RE = re.compile(r'(ohje(et)?|pisteet)(\s|$)', re.IGNORECASE)

def _is_helper_sheet(name: str) -> bool:
    n = name.lower().strip()
    if n == "data" or n.startswith(("data-", "data ", "data_")):
        return True
    return bool(_HELPER_RE.match(n))


# ── requirement row detection ─────────────────────────────────────────────────

# Matches "A 1", "A1", "A 10", "D 1", "10", "1", and sub-numbered ids like "1.1", "1-2",
# "7.12", "2-3" — letter prefix is optional. Sub-number parts are capped at 3 digits so genuine
# requirement ids match but date/year fragments (e.g. "1.2.2024-18.6.2024") do not. The $ anchor
# prevents matching long instruction texts that start with a letter+digit. A value with no
# separator reduces to the original pattern, so single-numbered rows match exactly as before.
_REQ_ROW_RE = re.compile(r'^[A-Z]{0,2}\s*\d+([.\-]\d{1,3})*\s*$', re.IGNORECASE)

def _is_req_row(b_val) -> bool:
    if b_val is None:
        return False
    return bool(_REQ_ROW_RE.match(str(b_val).strip()))


# ── format detection ──────────────────────────────────────────────────────────

_DATE_PREFIX = re.compile(r'^\d{1,2}/\d{4}')

def _is_scoring(d) -> bool:
    """True if the requirement-type tag marks a scored (vs mandatory) row. Keyed on the
    stem 'pisteyt' so inflections (pisteytettävä, pisteytetään, pisteytys, …) all match.
    'pisteet' (the score column) has no 'y' and is not matched."""
    return bool(d and "pisteyt" in str(d).lower())


def _detect_format(ws) -> int:
    """
    0 = unknown — no requirement rows and not a Format-4 table (nothing to extract; skipped)
    1 = narrative (G starts with date like "08/2015")
    2 = parallel columns (G and H are both numbered lists in separate columns)
    3 = consolidated prose (fallback — G contains plain prose, not a date or numbered list)
    4 = row-per-project table (each project is its own row, with a labelled column header
        row). Identified by a "Nro" header row that has several headed columns in G..N.
    Inspects only mandatory requirement rows for detection (1/2/3).
    """
    # Single pass over the sheet. Re-iterating an openpyxl read-only worksheet can
    # mis-read cells, so Format-4 detection is folded into this same loop rather than
    # done in a separate pass.
    saw_req_row = False
    for row in ws.iter_rows(min_row=2, values_only=True):
        b = row[1] if len(row) > 1 else None
        c = row[2] if len(row) > 2 else None
        d = row[3] if len(row) > 3 else None
        g = str(row[6]).strip() if len(row) > 6 and row[6] else ""
        h = str(row[7]).strip() if len(row) > 7 and row[7] else ""

        # Format 4 (row-per-project) is a column header row ("Nro" in B) with several
        # headed project-attribute columns in G..N. Genuine Format 1/2/3 sheets have at
        # most 1-2 headed columns there; row-per-project tables have 5+, so the threshold
        # cleanly separates them and is template-agnostic (any client's column names work).
        if str(b or "").strip() == "Nro" and sum(1 for i in range(6, 14) if _s(row, i)) >= _FMT4_MIN_HEADED:
            return 4

        if not _is_req_row(b):
            continue
        saw_req_row = True
        if not c or not g:
            continue
        if _is_scoring(d):
            continue  # skip scoring rows for format detection

        if _DATE_PREFIX.match(g):
            return 1
        if re.match(r'^\d+\.\s*\S', g) and re.match(r'^\d+\.\s*\S', h):
            return 2
        # keep scanning — a plain-text row (e.g. "Koulutus") must not
        # short-circuit detection before the parallel-column rows are seen

    # A sheet with requirement rows but no positive 1/2/4 signal is consolidated prose
    # (Format 3). A sheet with no requirement rows at all is not an extractable expert sheet
    # (cover page, summary, unrecognised layout) — mark it unknown so the caller skips it
    # explicitly rather than running the Format-3 path over it for 0 rows.
    return 3 if saw_req_row else 0


# ── name / role finder ────────────────────────────────────────────────────────

# Matches "Asiantuntijan rooli:", "Asiantuntijan 1 rooli:", "Asiantuntijan 1. rooli:"
_ROLE_MARKER = re.compile(r'asiantuntijan[\s\d.]*rooli\s*:\s*', re.IGNORECASE)
# Matches a "Asiantuntijan nimi:" label cell (name lives in an adjacent cell, not here).
_NAME_LABEL = re.compile(r'^asiantuntijan\s+nimi\s*:?\s*$', re.IGNORECASE)


def _looks_like_name(v: str) -> bool:
    """Heuristic: 2-4 capitalised whitespace-separated tokens, no digits, not a template
    label. Used only for structurally-anchored cells (a header row above the requirement
    table), so the loose shape is acceptable."""
    v = v.strip()
    if not v or _is_fake_name(v) or any(ch.isdigit() for ch in v):
        return False
    toks = v.split()
    return 2 <= len(toks) <= 4 and all(t[:1].isupper() for t in toks) and len(v) <= 40


def _find_name_role(ws) -> tuple[str, str]:
    """Resolve (developer_name, role) for the sheet, trying three conventions in order:

    1. an explicit "Asiantuntijan rooli:" marker in col B with the name in col D;
    2. an "Asiantuntijan nimi:" label in col B with the name in the next non-empty cell;
    3. role-as-sheet-name layouts: a header row above the requirement table whose col B is a
       role label and whose col D or E holds a name-like value.

    Patterns 2/3 are bounded to the rows before the first requirement row (where headers
    live) and never override pattern 1, so files that already match pattern 1 are unchanged.
    """
    rows = list(ws.iter_rows(min_row=1, max_row=50, values_only=True))
    title = ws.title.strip()

    # Pattern 1 — explicit role marker (unchanged behaviour).
    for row in rows:
        b = _s(row, 1)
        d = _s(row, 3)
        if _ROLE_MARKER.search(b) and d:
            return d, _ROLE_MARKER.sub("", b).strip().rstrip(":")

    # Only scan the header area (above the requirement table) for patterns 2/3.
    first_req = next((i for i, r in enumerate(rows)
                      if _is_req_row(r[1] if len(r) > 1 else None)), len(rows))
    header_rows = rows[:first_req]

    # Pattern 2 — "Asiantuntijan nimi:" label with the name in an adjacent cell.
    for row in header_rows:
        if _NAME_LABEL.match(_s(row, 1)):
            for ci in (2, 3, 4):  # C, D, E
                v = _s(row, ci)
                if v and not _is_fake_name(v):
                    return v, title

    # Pattern 3 — role in col B, name-like value in col D or E (role-as-sheet-name).
    for row in header_rows:
        b = _s(row, 1)
        if not b or b == "Nro" or b.lower().startswith(("ohje", "tarjoajan nimi")):
            continue
        # B must be a role label, not itself a person name — otherwise this is a data row in a
        # comparison/listing sheet, not a header, and would yield a bogus name/role pair.
        if _looks_like_name(b):
            continue
        for ci in (3, 4):  # D, E
            v = _s(row, ci)
            if _looks_like_name(v):
                role = b.split("\n")[0].strip().rstrip(":") or title
                return v, role

    return "", ""


# ── template noise detection ──────────────────────────────────────────────────

_NOISE = (
    # Column header / sub-header instructions
    "kuvaus", "asiakkaat", "toimeksiantojen", "kuvaile", "kirjoita",
    "vastauksesta tulee ilmetä", "tarkennukset tulee", "esimerkkikuvaus",
    # Unfilled template instruction cells ("Ohje tarjoajalle: Täytä...")
    "ohje tarjoajalle",
    # Unfilled "fill in here" placeholders
    "täytä kuvaus", "täytä tähän", "täytä kokemus",
    # Template example answers left in place
    "esimerkki vastauksesta", "esimerkkiasiakas",
)

def _is_template(val: str) -> bool:
    v = val.lower().strip()
    return any(v.startswith(t) for t in _NOISE)


# ── fake / unfilled developer name detection ──────────────────────────────────

# Placeholder / non-person values that unfilled templates leave in the name cell. A real expert
# name never starts with these tokens, ends with a company suffix, or contains a digit / one of
# the structural characters below. Rejecting them keeps template sheets from being treated as
# real experts — so the M4 "unclassified sheet" warning fires only for genuinely new *filled*
# layouts, not for the ~100 blank templates in the corpus (see [[m5-multi-expert-per-sheet]]).
_FAKE_NAME_PREFIX = (
    "asiantuntijan nimi",   # header label left unfilled ("Asiantuntijan nimi (nimettävä...)")
    "ohje tarjoajalle",     # template instruction left in the name cell
    "etunimi", "sukunimi",  # "Etunimi Sukunimi" placeholder (AIPA, KEHA templates)
    "valitse alasvetovalikosta",  # unfilled dropdown (ORK Power Platform)
    "täytä",                # "(Täytä: etunimi sukunimi)" CV-lomake placeholder
    "ks. vaatimus", "ks vaatimus",  # "see requirement" (Metropolia team cards)
    "n.n", "n. n",          # N.N placeholder
    "nimi:",                # bare "Nimi:" label
    "vastaus",              # column header "Vastaus vähimmäisvaatimukseen" mis-read as a name
    "tarjoaja",             # instruction "Tarjoaja valitsee alasvetovalikosta ..." / "Tarjoajan nimi:"
)
# Substrings that only appear in template example/instruction values, never a real name.
_FAKE_NAME_SUBSTR = ("esimerkki",)  # "Esko Esimerkki" = "Esko Example"
# A trailing "nimi" is a leftover label ("Projektipäällikön nimi"), never a real surname. NB: a
# company suffix (" oy"/" ab") is deliberately NOT rejected — the bidder company legitimately
# appears in the name cell on some filled sheets (e.g. Fintraffic), carrying real experience rows.
_FAKE_NAME_SUFFIX = ("nimi",)
# Characters that never occur in a person name but do in template placeholders ("Hinta (€/htp)").
# '/' is intentionally excluded: real cells list several people as "Sanna/Reko/Antti".
_FAKE_NAME_CHARS = set("0123456789(€@")

def _is_fake_name(val: str) -> bool:
    v = (val or "").lower().strip().lstrip("(").strip()
    if not v:
        return True
    if any(v.startswith(t) for t in _FAKE_NAME_PREFIX):
        return True
    if any(v.endswith(t) for t in _FAKE_NAME_SUFFIX):
        return True
    if any(s in v for s in _FAKE_NAME_SUBSTR):
        return True
    return any(c in _FAKE_NAME_CHARS for c in v)


# ── cell value helper ─────────────────────────────────────────────────────────

def _s(row, idx: int) -> str:
    if idx >= len(row) or row[idx] is None:
        return ""
    return str(row[idx]).strip()


# ── evidence builders ─────────────────────────────────────────────────────────

def _evidence_fmt1_fmt3(row, is_scoring: bool) -> str:
    """Format 1 and 3 fallback: G = mandatory evidence, H = scoring evidence. Used only when a
    sheet has no recognisable evidence-header row (see _fmt13_evidence_header)."""
    return _s(row, 7 if is_scoring else 6)


def _fmt13_evidence_header(row) -> tuple:
    """If `row` is a Format 1/3 section header, return (evidence_col_index, is_scoring);
    otherwise None.

    The evidence column is the one headed "Kuvaus ... täyttymisestä" (mandatory section) or
    "Kuvaus pisteytettäv..." (scoring section). Its position varies by template (seen in F, G
    and H), so it is located by header text instead of being assumed at G/H. The scoring test
    is checked first because a scoring header can also contain "täyttymisestä"."""
    for i in range(2, 14):  # scan C..N
        s = _s(row, i).split("\n")[0].strip().lower()
        if not s:
            continue
        if s.startswith("kuvaus pisteyt"):
            return (i, True)
        if s.startswith("kuvaus") and "täyttymisest" in s:
            return (i, False)
    return None


# Fallback Format-2 column labels (source order G..L) when a sheet has no sub-header row.
_FMT2_DEFAULT = [(6, "Asiakkaat"), (7, "Projektit"), (8, "Ajankohta"),
                 (9, "Rooli"), (10, "HTP"), (11, "Kuvaus")]

def _fmt2_header_map(row) -> list:
    """[(col_index, label)] for the headed G..N columns of a Format-2 sub-header row.
    The label is the header's first line (the rest is filler like "Vastauksesta tulee...").
    Mandatory sections head columns from G; scoring sections from H (G is the score)."""
    cols = [(i, _s(row, i).split("\n")[0].strip()) for i in range(6, 14) if _s(row, i)]
    return cols if len(cols) >= 2 else []

def _evidence_fmt2(row, header_cols) -> str:
    """
    Format 2: parallel numbered-list columns. Stored verbatim — never split or zipped —
    so unequal column lengths cannot silently drop or mis-pair items. Columns are labelled
    by the sheet's own sub-header row (header_cols), so different templates' column sets and
    orders are handled; headerless columns are skipped.
    """
    first_idx = header_cols[0][0]

    # Genuine parallel numbered lists → store each headed column verbatim, labelled.
    if re.match(r'^\d+\.\s', _s(row, first_idx)):
        blocks = []
        for idx, label in header_cols:
            val = _s(row, idx)
            if val:
                blocks.append(f"{label}:\n{val}")
        return "\n\n".join(blocks)

    # Free-text row inside a Format-2 sheet (e.g. "Koulutus", a certification answer).
    # Not a column structure → join non-empty headed cells verbatim, no labels.
    parts = [_s(row, idx) for idx, _ in header_cols]
    val = "\n".join(p for p in parts if p)
    return "" if _is_template(val) else val


# Format 4 (row-per-project) is header-driven: each project is a row, and evidence is
# built from the columns that have a header, labelled by that header. Columns with no
# header (e.g. a client's "Projekti 1" placeholder column) are ignored. The minimum number
# of headed G..N columns that marks a sheet as a project table (vs a 1-2 column Format 1/2/3
# sheet) — observed data shows a clean gap (Format 1/2/3 ≤ 2, project tables ≥ 5).
_FMT4_MIN_HEADED = 4

def _fmt4_header_cols(rows) -> list:
    """[(col_index, label)] for the headed G..N columns of the 'Nro' header row."""
    for row in rows:
        if str((row[1] if len(row) > 1 else "") or "").strip() == "Nro":
            cols = []
            for i in range(6, 14):
                label = _s(row, i)
                if label:
                    cols.append((i, re.sub(r"\s+", " ", label).strip()))
            if len(cols) >= _FMT4_MIN_HEADED:
                return cols
    return []

def _is_fmt4_project_row(row, header_cols) -> bool:
    """A project/continuation row: not a requirement row, not a repeated header row,
    and carrying a value in at least one headed column."""
    if _is_req_row(row[1] if len(row) > 1 else None):
        return False
    if str((row[1] if len(row) > 1 else "") or "").strip() == "Nro":
        return False
    return any(_s(row, idx) for idx, _ in header_cols)

def _evidence_fmt4(project_rows, header_cols) -> str:
    """One labelled block per project row, values verbatim, empty cells omitted.
    Labels are the source column headers; headerless columns are not included."""
    blocks = []
    for row in project_rows:
        lines = []
        for idx, label in header_cols:
            val = _s(row, idx)
            if val:
                lines.append(f"{label}: {val}")
        if lines:
            blocks.append("\n".join(lines))
    return "\n\n".join(blocks)


# ── per-sheet extraction ──────────────────────────────────────────────────────

def _extract_sheet(ws, rel_path: str, file_name: str,
                   mtime: str, today: str) -> list[dict]:
    developer_name, role = _find_name_role(ws)
    if _is_fake_name(developer_name):
        return []
    fmt = _detect_format(ws)
    if fmt == 0:
        # A real expert name resolved but the sheet has no requirement rows to extract from.
        # Skip explicitly and warn (file+sheet only, no name) so a genuine miss is visible
        # instead of silently producing 0 rows via the Format-3 fallback.
        print(f"WARNING: unclassified sheet skipped (no requirement rows): "
              f"{rel_path} :: {ws.title}", file=sys.stderr)
        return []
    if fmt == 4:
        return _extract_sheet_fmt4(ws, developer_name, role,
                                   rel_path, file_name, mtime, today)
    if fmt == 2:
        return _extract_sheet_fmt2(ws, developer_name, role,
                                   rel_path, file_name, mtime, today)

    # Format 1/3. A sheet can stack several experts, each opening with an "Asiantuntijan
    # rooli:" marker (B = role, D = name) and carrying its own header rows + requirement rows.
    # With ≥2 markers, split into per-expert blocks so each requirement is attributed to the
    # right person; otherwise it is one expert over the whole sheet (the common case, unchanged).
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    markers = [i for i, r in enumerate(rows)
               if _ROLE_MARKER.search(_s(r, 1)) and _s(r, 3)]
    if len(markers) >= 2:
        records = []
        for bi, start in enumerate(markers):
            end = markers[bi + 1] if bi + 1 < len(markers) else len(rows)
            mrow = rows[start]
            bname = _s(mrow, 3)
            if _is_fake_name(bname):
                continue
            # Role is usually appended after the marker in B; some templates (e.g. PRH) put the
            # marker label alone in B and the role text in C, so fall back to C when B is bare.
            brole = _ROLE_MARKER.sub("", _s(mrow, 1)).strip().rstrip(":") or _s(mrow, 2)
            records += _extract_fmt13_rows(rows[start:end], bname, brole,
                                           rel_path, file_name, ws.title, mtime, today)
        return records
    return _extract_fmt13_rows(rows, developer_name, role,
                               rel_path, file_name, ws.title, mtime, today)


def _extract_fmt13_rows(rows, developer_name: str, role: str, rel_path: str,
                        file_name: str, sheet_title: str, mtime: str, today: str) -> list[dict]:
    """Extract Format 1/3 records for ONE expert from a list of rows.

    Evidence is read from the cell under the section's "Kuvaus ..." header, tracked forward as
    section headers are passed (mandatory section, then scoring section), because its position
    varies by template (F/G/H). A row range with no recognisable header falls back to the fixed
    G (mandatory) / H (scoring) columns.
    """
    records = []
    cur_ev_col = None
    for row in rows:
        if len(row) < 4:
            continue

        b, c, d = row[1], row[2], row[3]

        if not _is_req_row(b):
            # Non-requirement row: may be a section header that relocates the evidence column.
            hdr = _fmt13_evidence_header(row)
            if hdr is not None:
                cur_ev_col = hdr[0]
            continue

        req_text = str(c or "").strip()
        if not req_text or _is_template(req_text):
            continue

        if cur_ev_col is not None:
            evidence = _s(row, cur_ev_col)
        else:
            evidence = _evidence_fmt1_fmt3(row, _is_scoring(d))
        if not evidence or _is_template(evidence):
            continue

        records.append({
            "developer_name": developer_name,
            "role": role,
            "requirement_text": req_text,
            "evidence": evidence,
            "technologies": "",
            "domain_or_industry": "",
            "source_file_name": file_name,
            "source_relative_path": rel_path,
            "source_sheet": sheet_title,
            "source_last_modified": mtime,
            "extracted_date": today,
        })

    return records


def _extract_sheet_fmt2(ws, developer_name: str, role: str, rel_path: str,
                        file_name: str, mtime: str, today: str) -> list[dict]:
    """Format 2: parallel numbered-list columns, labelled by the sheet's own sub-header
    row. Each requirement section has a "Nro" row followed by a sub-header row whose headed
    columns name the lists; requirement rows in that section are labelled from it. This
    handles templates with different column sets/orders (and the per-section scoring shift,
    where the sub-header starts at H because G holds the score)."""
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    records = []
    headers = _FMT2_DEFAULT          # used until the first sub-header row is seen
    prev_was_nro = False
    for row in rows:
        b = row[1] if len(row) > 1 else None
        b_str = str(b or "").strip()

        # The row right after a "Nro" row is the per-section sub-header row.
        if prev_was_nro:
            prev_was_nro = False
            if not _is_req_row(b):
                hm = _fmt2_header_map(row)
                if hm:
                    headers = hm
                continue  # sub-header row is not a requirement
        if b_str == "Nro":
            prev_was_nro = True
            continue

        if not _is_req_row(b):
            continue
        req_text = str(row[2] or "").strip()  # col C
        if not req_text or _is_template(req_text):
            continue

        # Labelled evidence starts with a header like "Asiakkaat:"; the empty check is the
        # only gate here (template filtering happens in _evidence_fmt2's free-text branch).
        evidence = _evidence_fmt2(row, headers)
        if not evidence:
            continue

        records.append({
            "developer_name": developer_name,
            "role": role,
            "requirement_text": req_text,
            "evidence": evidence,
            "technologies": "",
            "domain_or_industry": "",
            "source_file_name": file_name,
            "source_relative_path": rel_path,
            "source_sheet": ws.title,
            "source_last_modified": mtime,
            "extracted_date": today,
        })

    return records


def _extract_sheet_fmt4(ws, developer_name: str, role: str, rel_path: str,
                        file_name: str, mtime: str, today: str) -> list[dict]:
    """Format 4: each requirement's evidence spans several rows (one per project).

    The requirement row carries the first project; following continuation rows carry the
    rest. Evidence is built from the headed columns (labelled by the source header row),
    so headerless placeholder columns (e.g. a client's "Projekti 1") are ignored.
    """
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    header_cols = _fmt4_header_cols(rows)
    records = []
    i, n = 0, len(rows)
    while i < n:
        row = rows[i]
        if not _is_req_row(row[1] if len(row) > 1 else None):
            i += 1
            continue

        req_text = str(row[2] or "").strip()  # col C
        if not req_text or _is_template(req_text):
            i += 1
            continue

        project_rows = [row]  # the requirement row carries the first project
        j = i + 1
        while j < n and _is_fmt4_project_row(rows[j], header_cols):
            project_rows.append(rows[j])
            j += 1

        evidence = _evidence_fmt4(project_rows, header_cols)
        if evidence:
            records.append({
                "developer_name": developer_name,
                "role": role,
                "requirement_text": req_text,
                "evidence": evidence,
                "technologies": "",
                "domain_or_industry": "",
                "source_file_name": file_name,
                "source_relative_path": rel_path,
                "source_sheet": ws.title,
                "source_last_modified": mtime,
                "extracted_date": today,
            })
        i = j

    return records


# ── per-file extraction ───────────────────────────────────────────────────────

def extract_file(path: Path) -> list[dict]:
    rel = unicodedata.normalize("NFC", str(path.relative_to(SYNC_ROOT)))
    try:
        mtime = datetime.fromtimestamp(path.stat().st_mtime).strftime("%Y-%m-%d %H:%M")
    except OSError:
        mtime = ""
    today = date.today().isoformat()

    # Let an open failure propagate: a locked or corrupt file is NOT "0 records" — callers
    # (run.py sync) must be able to tell the difference so they retry it next run instead of
    # marking it synced. A genuinely empty sheet still returns [] below.
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)

    records = []
    for ws in wb.worksheets:
        if _is_helper_sheet(ws.title):
            continue
        records.extend(_extract_sheet(ws, rel, path.name, mtime, today))

    wb.close()
    # Drop rows whose evidence merely echoes the requirement text — that is a non-answer
    # (seen when a layout has no real evidence column), not experience. Real evidence never
    # equals the requirement verbatim.
    return [r for r in records
            if r["evidence"].strip().lower() != r["requirement_text"].strip().lower()]


# ── CLI ───────────────────────────────────────────────────────────────────────

def main():
    args = sys.argv[1:]
    if not args:
        print(
            "Usage:\n"
            "  python extract_requirements.py <file1> [<file2> ...]\n"
            "  python extract_requirements.py --all",
            file=sys.stderr,
        )
        sys.exit(1)

    if args == ["--all"]:
        candidates, dropped = deduplicate(enumerate_candidates())
        print(f"Candidates: {len(candidates)}  ({len(dropped)} dropped by file-level dedupe)")
        paths = candidates
    else:
        paths = []
        for a in args:
            p = Path(a)
            if not p.is_absolute():
                p = SYNC_ROOT / a
            paths.append(p)

    all_records: list[dict] = []
    for i, path in enumerate(paths, 1):
        try:
            label = str(path.relative_to(SYNC_ROOT))
        except ValueError:
            label = path.name
        print(f"[{i}/{len(paths)}] {label} ... ", end="", flush=True)
        try:
            recs = extract_file(path)
        except Exception as e:
            print(f"ERROR: {e}")
            continue
        print(f"{len(recs)} records")
        all_records.extend(recs)

    OUTPUT_FILE.write_text(
        json.dumps(all_records, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    missing_text = sum(1 for r in all_records if not r.get("requirement_text"))
    empty_ev    = sum(1 for r in all_records if not r.get("evidence"))
    print(f"\nTotal records:      {len(all_records)}")
    print(f"Missing req_text:   {missing_text}")
    print(f"Empty evidence:     {empty_ev}")
    print(f"Written to:         {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
