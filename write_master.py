"""Write the master Excel file from extracted records (v2 requirement-row schema)."""
import hashlib
import json
import re
import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from config import SYNC_ROOT, MASTER_PATH

# ── column definition (spec section 6) ───────────────────────────────────────

COLUMNS = [
    "developer_name",
    "role",
    "requirement_text",
    "evidence",
    "technologies",
    "domain_or_industry",
    "source_file_name",
    "source_relative_path",
    "source_sheet",
    "source_last_modified",
    "extracted_date",
]

COLUMN_WIDTHS = {
    "developer_name": 22,
    "role": 30,
    "requirement_text": 60,
    "evidence": 100,
    "technologies": 40,
    "domain_or_industry": 28,
    "source_file_name": 40,
    "source_relative_path": 55,
    "source_sheet": 20,
    "source_last_modified": 20,
    "extracted_date": 14,
}

HEADER_LABELS = {
    "developer_name": "Asiantuntijan nimi",
    "role": "Rooli tarjouksessa",
    "requirement_text": "Vaatimus / osaaminen",
    "evidence": "Kokemus (alkuperäinen teksti)",
    "technologies": "Teknologiat",
    "domain_or_industry": "Toimiala",
    "source_file_name": "Lähdetiedosto",
    "source_relative_path": "Polku",
    "source_sheet": "Taulukko",
    "source_last_modified": "Tiedosto muokattu",
    "extracted_date": "Poimittu",
}


# ── dedupe key ────────────────────────────────────────────────────────────────

def dedupe_key(rec: dict) -> tuple:
    # Identity is the content itself: (developer, requirement, evidence). The master collects
    # UNIQUE expert experience, so the same fact appearing in several source files (a draft and
    # its final submission, or one expert proposed across tenders) is one row, not many — the
    # source path/sheet are provenance, not identity, and are deliberately NOT in the key.
    # evidence stays in the key so a mandatory row and a scoring row that share a requirement_text
    # but carry different evidence remain the two distinct rows they are (the R4 fix); only
    # byte-identical evidence collapses. developer_name already separates different experts.
    return (
        (rec.get("developer_name") or "").strip().lower(),
        (rec.get("requirement_text") or "").strip().lower(),
        (rec.get("evidence") or "").strip(),
    )


# Backwards-compatible private alias (kept so existing internal refs keep working).
_dedupe_key = dedupe_key


def content_hash(rec: dict) -> str:
    """Stable string id for a record's CONTENT, used as the key of the enrichment side table
    (`enrichment.json`). It hashes the exact `dedupe_key` tuple, so an enrichment entry maps 1:1
    to a deduped master row and re-attaches after re-extraction as long as the content is
    unchanged. NUL-joined because none of the fields can contain NUL, so the join is unambiguous."""
    return hashlib.sha1("\x00".join(dedupe_key(rec)).encode("utf-8")).hexdigest()


def apply_enrichment(records: list[dict], enrichment: dict) -> int:
    """Fill `technologies` / `domain_or_industry` on each record from the content-keyed side
    table (authoritative for these two fields). Returns the number of records that received at
    least one non-empty tag. Called at master-rebuild time so enrichment lives outside the batch
    and survives re-sync. Records absent from the table are left untouched."""
    touched = 0
    for rec in records:
        entry = enrichment.get(content_hash(rec))
        if not entry:
            continue
        tech = (entry.get("technologies") or "").strip()
        dom = (entry.get("domain_or_industry") or "").strip()
        if tech:
            rec["technologies"] = tech
        if dom:
            rec["domain_or_industry"] = dom
        if tech or dom:
            touched += 1
    return touched


# ── bookkeeping: resolve mtime from disk ─────────────────────────────────────

def _source_mtime(source_relative_path: str) -> str:
    p = SYNC_ROOT / source_relative_path
    try:
        ts = p.stat().st_mtime
        return datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M")
    except (FileNotFoundError, OSError):
        return ""


# ── load records from extraction JSON ────────────────────────────────────────

def load_records(json_path: Path) -> list[dict]:
    data = json.loads(json_path.read_text(encoding="utf-8"))
    raw = data.get("records", data) if isinstance(data, dict) else data
    records = []
    today = date.today().isoformat()
    for rec in raw:
        row = {col: "" for col in COLUMNS}
        for col in COLUMNS:
            val = rec.get(col, "")
            # requirement_text and evidence are verbatim source text — never transform
            if col not in ("requirement_text", "evidence") and isinstance(val, list):
                val = "; ".join(str(v) for v in val)
            row[col] = val if val is not None else ""
        if not row["source_last_modified"] and row["source_relative_path"]:
            row["source_last_modified"] = _source_mtime(row["source_relative_path"])
        if not row["extracted_date"]:
            row["extracted_date"] = today
        records.append(row)
    return records


# ── row-level dedupe ──────────────────────────────────────────────────────────

def dedupe(records: list[dict]) -> list[dict]:
    """Keep newest mtime per (developer_name, requirement_text, evidence) — see _dedupe_key."""
    groups: dict[tuple, list[dict]] = {}
    for rec in records:
        key = _dedupe_key(rec)
        groups.setdefault(key, []).append(rec)
    result = []
    for group in groups.values():
        if len(group) == 1:
            result.append(group[0])
        else:
            winner = max(group, key=lambda r: r.get("source_last_modified", ""))
            result.append(winner)
    return result


# ── Excel write ───────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
CELL_FONT   = Font(name="Calibri", size=10)
WRAP_ALIGN  = Alignment(wrap_text=True, vertical="top")


def write_excel(records: list[dict], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Asiantuntijat"

    # Header row
    for col_idx, col_key in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=HEADER_LABELS[col_key])
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 28

    # Data rows
    for row_idx, rec in enumerate(records, start=2):
        for col_idx, col_key in enumerate(COLUMNS, start=1):
            val = rec.get(col_key)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = CELL_FONT
            cell.alignment = WRAP_ALIGN
        ws.row_dimensions[row_idx].height = 60

    # Column widths
    for col_idx, col_key in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS[col_key]

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}1"

    wb.save(path)
